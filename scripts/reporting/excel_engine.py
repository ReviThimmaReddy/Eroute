import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
from datetime import datetime

class ExcelReportGenerator:
    def __init__(self, title, app_name):
        self.wb = openpyxl.Workbook()
        self.title = title
        self.app_name = app_name
        self.timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def apply_header_style(self, cell):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2D3436", end_color="2D3436", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def create_summary_sheet(self, summary_data):
        sheet = self.wb.active
        sheet.title = "Summary"

        # Title
        sheet.merge_cells('A1:E2')
        cell = sheet['A1']
        cell.value = f"{self.title}"
        cell.font = Font(size=18, bold=True)
        cell.alignment = Alignment(horizontal="center")

        sheet['A4'] = "Application Name:"
        sheet['B4'] = self.app_name
        sheet['A5'] = "Execution Date:"
        sheet['B5'] = self.timestamp

        # KPI Cards
        headers = ["Total Tests", "Passed", "Failed", "Skipped", "Pass Rate"]
        values = [
            summary_data['total'],
            summary_data['passed'],
            summary_data['failed'],
            summary_data['skipped'],
            f"{summary_data['pass_rate']}%"
        ]

        for i, header in enumerate(headers):
            cell = sheet.cell(row=7, column=i+1)
            cell.value = header
            self.apply_header_style(cell)

            val_cell = sheet.cell(row=8, column=i+1)
            val_cell.value = values[i]
            val_cell.alignment = Alignment(horizontal="center")
            if header == "Passed" and summary_data['passed'] > 0:
                val_cell.font = Font(color="00B894", bold=True)
            elif header == "Failed" and summary_data['failed'] > 0:
                val_cell.font = Font(color="D63031", bold=True)

    def create_results_sheet(self, results, columns):
        sheet = self.wb.create_sheet("Test Results")
        for col_idx, col_name in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.value = col_name
            self.apply_header_style(cell)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

        for row_idx, result in enumerate(results, 2):
            for col_idx, value in enumerate(result, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True)

                # Conditional Formatting
                if value == "PASSED":
                    cell.font = Font(color="00B894", bold=True)
                elif value == "FAILED":
                    cell.font = Font(color="D63031", bold=True)

    def save(self, path):
        self.wb.save(path)
